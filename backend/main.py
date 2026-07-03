from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Column configuration
THRESHOLD = 30
COL_FM50  = 14   # Column N — Overtime 50%
COL_FM100 = 15   # Column O — Overtime 100%
COL_LEAVE = 17   # Column Q — Paid Leave Entitlement (days)

# Cell styles
GREEN = PatternFill('solid', fgColor='FF92D050')  # Eligible employees
YELLOW = PatternFill('solid', fgColor='FFFFFF00') # Leave column header
WHITE = PatternFill('solid', fgColor='FFFFFFFF')  # Non-eligible rows

@app.post("/hesapla")
async def calculate(file: UploadFile = File(...)):
    # Save uploaded file to a temporary location
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path)
    ws = wb.active

    eligible = []
    total = 0

    # ── Sheet 1: Original data + leave entitlement column ──
    for row_idx in range(2, ws.max_row + 1):
        if not ws.cell(row_idx, 1).value:
            continue

        fm50  = ws.cell(row_idx, COL_FM50).value  or 0
        fm100 = ws.cell(row_idx, COL_FM100).value or 0

        # Write formula to leave entitlement column
        q = ws.cell(row_idx, COL_LEAVE)
        q.value = (
            f'=IF((N{row_idx}+O{row_idx})>{THRESHOLD},'
            f'ROUND(((N{row_idx}+O{row_idx}-{THRESHOLD})*1.5/9.5),1),"")'
        )
        q.alignment = Alignment(horizontal='center')
        total += 1

        if (fm50 + fm100) > THRESHOLD:
            # Highlight entire row green
            days = round(((fm50 + fm100 - THRESHOLD) * 1.5 / 9.5), 1)
            for c in range(1, ws.max_column + 1):
                ws.cell(row_idx, c).fill = GREEN
            eligible.append({
                'id':     ws.cell(row_idx, 1).value,
                'first':  ws.cell(row_idx, 2).value,
                'last':   ws.cell(row_idx, 3).value,
                'period': ws.cell(row_idx, 4).value,
                'fm50':   fm50,
                'fm100':  fm100,
                'total':  round(fm50 + fm100, 1),
                'days':   days
            })
        else:
            for c in range(1, ws.max_column + 1):
                ws.cell(row_idx, c).fill = WHITE

    # Style the leave column header
    ws.cell(1, COL_LEAVE).fill = YELLOW
    ws.cell(1, COL_LEAVE).font = Font(bold=True)
    ws.title = "All Employees"

    # ── Sheet 2: Summary ───────────────────────────────────
    summary = wb.create_sheet("Paid Leave Summary")
    thin = Side(style='thin', color='FF000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    summary.merge_cells('A1:H1')
    title = summary['A1']
    title.value = f"EMPLOYEES WITH PAID LEAVE ENTITLEMENT — {datetime.now().strftime('%B %Y')}"
    title.font = Font(bold=True, size=13, color='FFFFFFFF')
    title.fill = PatternFill('solid', fgColor='FF2E75B6')
    title.alignment = Alignment(horizontal='center', vertical='center')
    summary.row_dimensions[1].height = 28

    # Stats block
    summary['A3'] = 'Total Employees';      summary['B3'] = total
    summary['A4'] = 'Eligible for Leave';   summary['B4'] = len(eligible)
    summary['A5'] = 'Not Eligible';         summary['B5'] = total - len(eligible)
    for r in [3, 4, 5]:
        summary.cell(r, 1).font = Font(bold=True)
        summary.cell(r, 2).alignment = Alignment(horizontal='center')
    for c in [1, 2]:
        summary.cell(4, c).fill = PatternFill('solid', fgColor='FFE2EFDA')

    # Table headers
    headers = ['Employee ID', 'First Name', 'Last Name', 'Period',
               'OT 50%', 'OT 100%', 'Total OT', 'Leave Entitlement (Days)']
    summary.row_dimensions[7].height = 22
    for col, h in enumerate(headers, 1):
        c = summary.cell(7, col)
        c.value = h
        c.font = Font(bold=True, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    # Data rows
    for i, p in enumerate(eligible, 8):
        row_fill = PatternFill('solid', fgColor='FFE2EFDA' if i % 2 == 0 else 'FFFFFFFF')
        values = [p['id'], p['first'], p['last'], p['period'],
                  p['fm50'], p['fm100'], p['total'], p['days']]
        for col, v in enumerate(values, 1):
            c = summary.cell(i, col)
            c.value = v
            c.border = border
            c.fill = row_fill
            c.alignment = Alignment(horizontal='left' if col in [2, 3] else 'center')

    # Total row
    last_row = len(eligible) + 8
    summary.cell(last_row, 7).value = 'TOTAL'
    summary.cell(last_row, 7).font = Font(bold=True)
    summary.cell(last_row, 8).value = round(sum(p['days'] for p in eligible), 1)
    summary.cell(last_row, 8).font = Font(bold=True)
    for col in range(1, 9):
        summary.cell(last_row, col).fill = PatternFill('solid', fgColor='FFBDD7EE')
        summary.cell(last_row, col).border = border
    summary.cell(last_row, 8).alignment = Alignment(horizontal='center')

    # Column widths
    for i, width in enumerate([12, 14, 16, 12, 10, 10, 12, 22], 1):
        summary.column_dimensions[get_column_letter(i)].width = width

    # Save output file
    out_path = tmp_path.replace(".xlsx", "_result.xlsx")
    wb.save(out_path)
    os.unlink(tmp_path)

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"PaidLeave_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        headers={
            "X-Toplam": str(total),
            "X-Izin-Alan": str(len(eligible)),
            "Access-Control-Expose-Headers": "X-Toplam, X-Izin-Alan"
        }
    )


@app.get("/")
def root():
    return {"message": "PDKS Paid Leave Calculator API is running ✅"}